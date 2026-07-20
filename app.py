"""
AACMH Kose Pending Order Report — Streamlit app.

Upload the daily orders CSV and the sale_order Excel export, and this app
will filter, extract, look up WMS status/payment, and build the pivot-style
Summary sheet — then let you download the finished report.
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---- filter values -----------------------------------------------------
J_VALS = ['ACCEPTED/PICKED', 'NEW', 'READY TO SHIP']
G_VALS = ['COMPLETED', 'PENDING']
BC_VALS = ['lazada', 'shopee']
BD_VALS = [
    'lazada-Lazada - Kose Singapore',
    'shopee-Shopee - Kose Singapore',
    'shopee-Decorté_SG',
    'lazada-Lazada - Decorte Singapore',
]

EXTRACT_COLS = {
    'order_number': 'Order Number (B)',
    'invoice_number': 'Invoice Number (D)',
    'payment_status': 'Payment Status (G)',
    'sku': 'SKU (H)',
    'custom_sku': 'Custom SKU (I)',
    'order_item_status': 'Order Item Status (J)',
    'item_title': 'Item Title (L)',
    'ordered_date': 'Ordered Date (AS)',
    'channel': 'Channel (BC)',
    'nickname': 'Nickname (BD)',
}

PALETTE = ['FF0000', '92D050', 'FFC000', '00B0F0', 'FFFF00']


# ---- pipeline steps ------------------------------------------------------

def clean_csv(uploaded_file) -> pd.DataFrame:
    df = pd.read_csv(uploaded_file, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].apply(
            lambda x: x.strip().replace('\r', '').replace('\n', '')
            if isinstance(x, str) else x
        )
    return df


def filter_and_extract(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in EXTRACT_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"CSV is missing expected columns: {missing}")

    mask = (
        df['order_item_status'].isin(J_VALS)
        & df['payment_status'].isin(G_VALS)
        & df['channel'].isin(BC_VALS)
        & df['nickname'].isin(BD_VALS)
    )
    filtered = df.loc[mask, list(EXTRACT_COLS.keys())].copy()
    filtered.columns = list(EXTRACT_COLS.values())
    filtered['WMS Status'] = ''
    filtered['WMS Payment'] = ''
    return filtered.reset_index(drop=True)


def apply_wms_lookup(ws, sale_order_file):
    sale = pd.read_excel(sale_order_file, dtype=str)
    for col in ('Channel Order ID', 'WMS Status', 'Payment Status'):
        if col not in sale.columns:
            raise ValueError(f"sale_order file is missing column: {col}")

    status_lookup = dict(zip(
        sale['Channel Order ID'].astype(str).str.strip(),
        sale['WMS Status'].astype(str).str.strip(),
    ))
    payment_lookup = dict(zip(
        sale['Channel Order ID'].astype(str).str.strip(),
        sale['Payment Status'].astype(str).str.strip(),
    ))

    headers = [c.value for c in ws[1]]
    order_col = headers.index('Order Number (B)') + 1
    status_col = headers.index('WMS Status') + 1
    payment_col = headers.index('WMS Payment') + 1

    matched = 0
    for row in range(2, ws.max_row + 1):
        order_id = str(ws.cell(row=row, column=order_col).value).strip()
        st_val = status_lookup.get(order_id)
        pay_val = payment_lookup.get(order_id)
        if st_val:
            ws.cell(row=row, column=status_col, value=st_val)
        if pay_val:
            ws.cell(row=row, column=payment_col, value=pay_val)
        if st_val or pay_val:
            matched += 1
    return matched


def strip_time_from_dates(ws):
    headers = [c.value for c in ws[1]]
    date_col = headers.index('Ordered Date (AS)') + 1
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=date_col).value
        if val:
            ws.cell(row=row, column=date_col, value=str(val).strip().split(' ')[0])


def format_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Arial', size=10, bold=(cell.row == 1))
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = 'A2'


def add_summary_sheet(wb, extract_df: pd.DataFrame):
    df = extract_df.copy()
    df['DateFmt'] = pd.to_datetime(
        df['Ordered Date (AS)'], format='%d/%m/%Y', errors='coerce'
    ).dt.strftime('%d-%b')
    df = df.dropna(subset=['DateFmt'])

    group_cols = ['Nickname (BD)', 'Order Item Status (J)', 'WMS Status']
    pt = df.groupby(group_cols + ['DateFmt']).size().unstack('DateFmt', fill_value=0)
    date_cols = sorted(pt.columns, key=lambda d: datetime.strptime(d, '%d-%b'))
    pt = pt[date_cols].reset_index().sort_values(group_cols).reset_index(drop=True)

    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    ws = wb.create_sheet('Summary', 0)

    headers = ['Channel', 'Status', 'WMS'] + date_cols
    header_fill = PatternFill('solid', start_color='305496')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    bold = Font(name='Arial', size=11, bold=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill, c.border, c.alignment = header_font, header_fill, border, center

    date_col_fills = {d: PALETTE[i % len(PALETTE)] for i, d in enumerate(date_cols)}
    n = len(pt)

    for i, row in pt.iterrows():
        r = i + 2
        for j, col_name in enumerate(group_cols, start=1):
            c = ws.cell(row=r, column=j, value=row[col_name])
            c.font, c.border, c.alignment = bold, border, center
        for k, d in enumerate(date_cols, start=4):
            val = row[d]
            c = ws.cell(row=r, column=k)
            c.border, c.alignment, c.font = border, center, bold
            if val and val > 0:
                c.value = int(val)
                c.fill = PatternFill('solid', start_color=date_col_fills[d])

    def merge_repeats(col_idx):
        start = 2
        for r in range(3, n + 3):
            cur = ws.cell(row=r, column=col_idx).value if r <= n + 1 else None
            prev = ws.cell(row=start, column=col_idx).value
            if r > n + 1 or cur != prev:
                if r - 1 > start:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=r - 1, end_column=col_idx)
                start = r

    merge_repeats(1)
    merge_repeats(2)

    total_row = n + 2
    ws.cell(row=total_row, column=1, value='Grand Total').font = bold
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=1).alignment = center
    for j in (1, 2, 3):
        ws.cell(row=total_row, column=j).border = border
    for k, d in enumerate(date_cols, start=4):
        c = ws.cell(row=total_row, column=k, value=int(pt[d].sum()))
        c.font, c.border, c.alignment = bold, border, center
        c.fill = PatternFill('solid', start_color='D9D9D9')

    widths = [30, 18, 20] + [10] * len(date_cols)
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build_report(csv_file, sale_order_file) -> bytes:
    df = clean_csv(csv_file)
    extract = filter_and_extract(df)

    buf = io.BytesIO()
    extract.to_excel(buf, index=False)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    matched = apply_wms_lookup(ws, sale_order_file)
    strip_time_from_dates(ws)
    format_sheet(ws)
    add_summary_sheet(wb, extract)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, len(extract), matched


# ---- Streamlit UI ---------------------------------------------------------

st.set_page_config(page_title="AACMH Kose Pending Order Report", page_icon="📦")
st.title("📦 AACMH Kose Pending Order Report")
st.write(
    "Upload the daily orders CSV and the sale_order Excel export. "
    "The app filters, extracts, matches WMS status/payment, and builds "
    "a color-coded pivot Summary sheet automatically."
)

csv_file = st.file_uploader("Orders CSV (e.g. ALL-08-Jul-2026.csv)", type=["csv"])
sale_order_file = st.file_uploader("Sale order Excel export", type=["xlsx"])
output_name = st.text_input("Output file name", value="AACMH Kose Pending Order Report.xlsx")

if st.button("Build report", type="primary", disabled=not (csv_file and sale_order_file)):
    try:
        with st.spinner("Building report..."):
            result_bytes, n_rows, n_matched = build_report(csv_file, sale_order_file)
        st.success(f"Done — {n_rows} rows extracted, {n_matched} matched to WMS data.")
        st.download_button(
            "⬇️ Download report",
            data=result_bytes,
            file_name=output_name if output_name.endswith(".xlsx") else output_name + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Something went wrong: {e}")
else:
    if not (csv_file and sale_order_file):
        st.info("Upload both files to enable the Build button.")

with st.expander("What this app does"):
    st.markdown(
        """
1. Converts the orders CSV to Excel.
2. Filters by Order Item Status (ACCEPTED/PICKED, NEW, READY TO SHIP),
   Payment Status (COMPLETED, PENDING), Channel (lazada, shopee), and
   Nickname (the 4 Kose storefronts).
3. Extracts columns B, D, G, H, I, J, L, AS, BC, BD.
4. Adds **WMS Status** / **WMS Payment**, looked up from the sale_order
   file by matching Order Number to Channel Order ID.
5. Strips the time off Ordered Date, keeping only the date.
6. Builds a pivot-style **Summary** sheet: rows = Nickname → Status → WMS
   (merged), columns = each date, values = order counts, with a colored
   Grand Total row.
        """
    )
